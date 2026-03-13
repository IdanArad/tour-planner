#!/usr/bin/env python3
"""
Import researched venue/festival/booking data from Excel files into SQL.

Reads the 3 Excel files in researched-data/ and generates SQL INSERT statements
for the discovered_venues and discovered_events tables.

Usage:
  python3 scripts/import-excel-data.py > supabase/import-researched-data.sql
  # Then run in Supabase SQL Editor or psql

Requires: pip3 install openpyxl
"""

import openpyxl
import json
import re
import sys
from datetime import datetime

# Comprehensive country normalization map
COUNTRY_MAP = {
    # Full names → ISO
    "Austria": "AT", "Belgium": "BE", "Bulgaria": "BG", "Canada": "CA",
    "Croatia": "HR", "Czechia": "CZ", "Czech Republic": "CZ",
    "Denmark": "DK", "Finland": "FI", "France": "FR", "Germany": "DE",
    "Greece": "GR", "Hungary": "HU", "Iceland": "IS", "Ireland": "IE",
    "Israel": "IL", "Italy": "IT", "Lithuania": "LT", "Latvia": "LV",
    "Luxembourg": "LU", "Malta": "MT", "Netherlands": "NL", "Norway": "NO",
    "Poland": "PL", "Portugal": "PT", "Romania": "RO", "Serbia": "RS",
    "Slovakia": "SK", "Slovenia": "SI", "Spain": "ES", "Sweden": "SE",
    "Switzerland": "CH", "UK": "GB", "United Kingdom": "GB",
    "USA": "US", "Argentina": "AR", "Australia": "AU", "Estonia": "EE",
    "Ukraine": "UA",
    # ISO codes → themselves
    "AT": "AT", "BE": "BE", "BG": "BG", "CA": "CA", "CH": "CH",
    "CZ": "CZ", "DE": "DE", "DK": "DK", "EE": "EE", "ES": "ES",
    "FI": "FI", "FR": "FR", "GB": "GB", "GR": "GR", "HR": "HR",
    "HU": "HU", "IE": "IE", "IL": "IL", "IS": "IS", "IT": "IT",
    "LT": "LT", "LU": "LU", "LV": "LV", "MT": "MT", "NL": "NL",
    "NO": "NO", "PL": "PL", "PT": "PT", "RO": "RO", "RS": "RS",
    "SE": "SE", "SI": "SI", "SK": "SK", "UA": "UA", "US": "US",
    "AR": "AR", "AU": "AU",
    # Messy values from the Excel data
    "USA Ohio": "US", "USA/ CA": "US", "USA?": "US", "US?": "US",
    "global based US": "US", "BE?": "BE", "Ukraine?": "UA",
    "SL": "SI",  # SL likely means Slovenia
    "switzerland": "CH",
}

# Values to skip (not a single country)
SKIP_COUNTRY = {
    "Europe", "EU", "WW", "worldwide", "CH EU WW", "EU / Worldwide",
    "EU UK Latin america", "EU\\uk", "uk\\eu", "FR/CH",
    "UK / USA / GR",
}


def normalize_country(raw):
    """Normalize a country value to ISO 2-letter code."""
    if not raw:
        return None
    s = str(raw).strip()
    if s in SKIP_COUNTRY:
        return None
    # Direct lookup
    if s in COUNTRY_MAP:
        return COUNTRY_MAP[s]
    # Try with "DE Hannover" → "DE"
    prefix = s.split()[0] if " " in s else s
    if prefix in COUNTRY_MAP:
        return COUNTRY_MAP[prefix]
    return None


def escape_sql(val):
    """Escape a string for SQL insertion."""
    if val is None:
        return "NULL"
    s = str(val).strip()
    if not s:
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def clean_email(val):
    """Extract the first valid-looking email from a cell."""
    if not val:
        return None
    s = str(val).strip()
    emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', s)
    return emails[0] if emails else None


def parse_capacity(val):
    """Parse capacity from various formats."""
    if val is None:
        return "NULL"
    if isinstance(val, (int, float)):
        return str(int(val))
    s = str(val).strip().replace(",", "")
    m = re.match(r'(\d+)', s)
    return m.group(1) if m else "NULL"


def import_venues():
    """Import venues from Venues.xlsx."""
    wb = openpyxl.load_workbook("researched-data/Venues.xlsx", read_only=True)
    venues = []
    idx = 0

    # Berlin sheet: Name, Size, Mail, Notes, summer 25, winter 26
    ws = wb["Berlin"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if len(row) < 2:
            continue
        name = row[0]
        size = row[1] if len(row) > 1 else None
        mail = row[2] if len(row) > 2 else None
        notes = row[3] if len(row) > 3 else None
        if not name or not str(name).strip():
            continue
        idx += 1
        venues.append({
            "source": "excel_import",
            "source_id": f"berlin_{idx}",
            "name": str(name).strip(),
            "city": "Berlin",
            "country": "DE",
            "capacity": size,
            "booking_email": clean_email(mail),
            "notes": notes,
        })

    # Germany sheet: Country, Winter 26, Summer 25, City, Venue, Link, Mail, Guarantee, Notes
    ws = wb["Germany"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if len(row) < 5:
            continue
        vals = list(row) + [None] * (9 - len(row))
        _, _, _, city, venue_name, link, mail, guarantee, notes = vals[:9]
        if not venue_name or not str(venue_name).strip():
            continue
        idx += 1
        venues.append({
            "source": "excel_import",
            "source_id": f"germany_{idx}",
            "name": str(venue_name).strip(),
            "city": str(city).strip() if city else None,
            "country": "DE",
            "capacity": None,
            "booking_email": clean_email(mail),
            "website_url": str(link).strip() if link and "http" in str(link) else None,
            "notes": notes,
        })

    # Europe sheet: Country, Summer 25, City, Venue, Link, Mail, Guarantee, Notes
    ws = wb["Europe"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if len(row) < 4:
            continue
        vals = list(row) + [None] * (8 - len(row))
        country, _, city, venue_name, link, mail, guarantee, notes = vals[:8]
        if not venue_name or not str(venue_name).strip():
            continue
        cc = normalize_country(country)
        idx += 1
        venues.append({
            "source": "excel_import",
            "source_id": f"europe_{idx}",
            "name": str(venue_name).strip(),
            "city": str(city).strip() if city else None,
            "country": cc,
            "capacity": None,
            "booking_email": clean_email(mail),
            "website_url": str(link).strip() if link and "http" in str(link) else None,
            "notes": notes,
        })

    # Israel sheet (offset header on row 2)
    ws = wb["Israel"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[2:]:
        if len(row) < 4:
            continue
        area, venue_name, link, guarantee = row[0], row[1], row[2], row[3] if len(row) > 3 else None
        capacity = row[4] if len(row) > 4 else None
        contact = row[5] if len(row) > 5 else None
        if not venue_name or not str(venue_name).strip():
            continue
        idx += 1
        venues.append({
            "source": "excel_import",
            "source_id": f"israel_{idx}",
            "name": str(venue_name).strip(),
            "city": str(area).strip() if area else "Israel",
            "country": "IL",
            "capacity": capacity,
            "booking_email": clean_email(contact),
            "website_url": str(link).strip() if link and "http" in str(link) else None,
        })

    wb.close()
    return venues


def import_festivals():
    """Import festivals from Festivals.xlsx (all sheets)."""
    wb = openpyxl.load_workbook("researched-data/Festivals.xlsx", read_only=True)
    events = []
    idx = 0

    # Sheets: 2025, 2026, 2027 all have same structure:
    # Country, Status, IG, Followup, Name, Dates, Web, Mail, Application dates, Notes
    # (2027 has Month instead of Followup at col 2)
    for year_sheet in ["2025", "2026", "2027"]:
        if year_sheet not in wb.sheetnames:
            continue
        ws = wb[year_sheet]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(c).lower() if c else "" for c in rows[0]]

        # Find column positions
        name_col = next((i for i, h in enumerate(headers) if h == "name"), 4)
        country_col = next((i for i, h in enumerate(headers) if h == "country"), 0)
        status_col = next((i for i, h in enumerate(headers) if h == "status"), 1)
        dates_col = next((i for i, h in enumerate(headers) if h == "dates"), 5)
        web_col = next((i for i, h in enumerate(headers) if h == "web"), 6)
        mail_col = next((i for i, h in enumerate(headers) if h == "mail"), 7)

        for row in rows[1:]:
            vals = list(row) + [None] * 12
            name = vals[name_col]
            if not name or not str(name).strip():
                continue
            country = vals[country_col]
            cc = normalize_country(country)
            web = vals[web_col]
            mail = vals[mail_col]
            dates = vals[dates_col]
            status = vals[status_col]
            idx += 1
            events.append({
                "source": f"excel_import",
                "source_id": f"fest{year_sheet}_{idx}",
                "name": str(name).strip(),
                "country": cc,
                "website_url": str(web).strip() if web and "http" in str(web) else None,
                "booking_email": clean_email(mail),
                "dates_raw": str(dates).strip() if dates else None,
                "status": str(status).strip() if status else None,
                "event_type": "festival",
            })

    # festivalticker sheet: Country, Status, IG, Name, City, Dates, Website, Mail, Notes
    if "List from festivalticker" in wb.sheetnames:
        ws = wb["List from festivalticker"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            vals = list(row) + [None] * 10
            country, status, _, name, city, dates, web, mail, notes = vals[:9]
            if not name or not str(name).strip():
                continue
            cc = normalize_country(country)
            idx += 1
            events.append({
                "source": "excel_festivalticker",
                "source_id": f"festivalticker_{idx}",
                "name": str(name).strip(),
                "city": str(city).strip() if city else None,
                "country": cc,
                "website_url": str(web).strip() if web and "http" in str(web) else None,
                "booking_email": clean_email(mail),
                "dates_raw": str(dates) if dates else None,
                "status": str(status).strip() if status else None,
                "event_type": "festival",
            })

    wb.close()
    return events


def import_bookings():
    """Import booking agencies/promoters from Bookings & Promoters.xlsx."""
    wb = openpyxl.load_workbook("researched-data/Bookings & Promoters .xlsx", read_only=True)
    contacts = []
    idx = 0

    # Germany sheet: City, Name, Website, Email, Phone, Mail Sent, Follow up, Tour Message, Notes
    ws = wb["Germany"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        vals = list(row) + [None] * 10
        city, name, website, email, phone = vals[:5]
        mail_sent = vals[5]
        notes = vals[8]
        if not name or not str(name).strip():
            continue
        idx += 1
        contacts.append({
            "source": "excel_import",
            "source_id": f"booking_de_{idx}",
            "name": str(name).strip(),
            "city": str(city).strip() if city else None,
            "country": "DE",
            "website_url": str(website).strip() if website and "http" in str(website) else None,
            "booking_email": clean_email(email),
            "phone": str(phone).strip() if phone else None,
            "contact_type": "booking_agency",
            "outreach_status": str(mail_sent).strip() if mail_sent else None,
            "notes": notes,
        })

    # Europe sheet: Country, Name, Website, Email, Phone, First mail, Follow up mail
    ws = wb["Europe"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        vals = list(row) + [None] * 10
        country, name, website, email, phone = vals[:5]
        mail_sent = vals[5]
        notes = vals[8]
        if not name or not str(name).strip():
            continue
        cc = normalize_country(country)
        idx += 1
        contacts.append({
            "source": "excel_import",
            "source_id": f"booking_eu_{idx}",
            "name": str(name).strip(),
            "country": cc,
            "website_url": str(website).strip() if website and "http" in str(website) else None,
            "booking_email": clean_email(email),
            "phone": str(phone).strip() if phone else None,
            "contact_type": "booking_agency",
            "outreach_status": str(mail_sent).strip() if mail_sent else None,
        })

    # USAOther sheet: Country, Name, Website, Email, Phone, First mail, Follow up mail
    if "USAOther" in wb.sheetnames:
        ws = wb["USAOther"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            vals = list(row) + [None] * 10
            country, name, website, email, phone = vals[:5]
            mail_sent = vals[5]
            if not name or not str(name).strip():
                continue
            cc = normalize_country(country)
            idx += 1
            contacts.append({
                "source": "excel_import",
                "source_id": f"booking_us_{idx}",
                "name": str(name).strip(),
                "country": cc or "US",
                "website_url": str(website).strip() if website and "http" in str(website) else None,
                "booking_email": clean_email(email),
                "phone": str(phone).strip() if phone else None,
                "contact_type": "booking_agency",
                "outreach_status": str(mail_sent).strip() if mail_sent else None,
            })

    # Support Slots sheet: Country, Name, Website, Email, Phone, First mail, Follow up mail
    if "Support Slots" in wb.sheetnames:
        ws = wb["Support Slots"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            vals = list(row) + [None] * 10
            country, name, website, email, phone = vals[:5]
            mail_sent = vals[5]
            if not name or not str(name).strip():
                continue
            cc = normalize_country(country)
            idx += 1
            contacts.append({
                "source": "excel_import",
                "source_id": f"booking_support_{idx}",
                "name": str(name).strip(),
                "country": cc,
                "website_url": str(website).strip() if website and "http" in str(website) else None,
                "booking_email": clean_email(email),
                "phone": str(phone).strip() if phone else None,
                "contact_type": "booking_agency",
                "outreach_status": str(mail_sent).strip() if mail_sent else None,
            })

    # Promoters sheet: Country, Name, Website, Email, Phone, Mail Sent, Follow up, Notes
    if "Promoters " in wb.sheetnames:
        ws = wb["Promoters "]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            vals = list(row) + [None] * 10
            country, name, website, email, phone = vals[:5]
            mail_sent = vals[5]
            notes = vals[7]
            if not name or not str(name).strip():
                continue
            cc = normalize_country(country)
            idx += 1
            contacts.append({
                "source": "excel_import",
                "source_id": f"promoter_{idx}",
                "name": str(name).strip(),
                "country": cc,
                "website_url": str(website).strip() if website and "http" in str(website) else None,
                "booking_email": clean_email(email),
                "phone": str(phone).strip() if phone else None,
                "contact_type": "promoter",
                "outreach_status": str(mail_sent).strip() if mail_sent else None,
                "notes": notes,
            })

    # Event Planners sheet
    if "Event Planners" in wb.sheetnames:
        ws = wb["Event Planners"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            vals = list(row) + [None] * 10
            country, name, website, email, phone = vals[:5]
            mail_sent = vals[5]
            if not name or not str(name).strip():
                continue
            cc = normalize_country(country)
            idx += 1
            contacts.append({
                "source": "excel_import",
                "source_id": f"planner_{idx}",
                "name": str(name).strip(),
                "country": cc,
                "website_url": str(website).strip() if website and "http" in str(website) else None,
                "booking_email": clean_email(email),
                "phone": str(phone).strip() if phone else None,
                "contact_type": "event_planner",
            })

    # Booking W festivals sheet (header in row 2): Country, Booking Agency, Link, Email, Phone, Email Status
    if "Booking W festivals" in wb.sheetnames:
        ws = wb["Booking W festivals"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[2:]:  # skip title row + header row
            vals = list(row) + [None] * 10
            country, name, link, email, phone, mail_status = vals[:6]
            if not name or not str(name).strip():
                continue
            cc = normalize_country(country)
            idx += 1
            contacts.append({
                "source": "excel_import",
                "source_id": f"booking_fest_{idx}",
                "name": str(name).strip(),
                "country": cc,
                "website_url": str(link).strip() if link and "http" in str(link) else None,
                "booking_email": clean_email(email),
                "phone": str(phone).strip() if phone else None,
                "contact_type": "booking_agency",
                "outreach_status": str(mail_status).strip() if mail_status else None,
            })

    wb.close()
    return contacts


def main():
    print("-- Generated by import-excel-data.py")
    print("-- Imports researched venue/festival/booking data into discovered_venues and discovered_events")
    print(f"-- Generated at: {datetime.now().isoformat()}")
    print()
    print("BEGIN;")
    print()

    # Clear previous imports
    print("DELETE FROM discovered_events WHERE source IN ('excel_import', 'excel_festivalticker');")
    print("DELETE FROM discovered_venues WHERE source = 'excel_import';")
    print()

    # Import venues
    venues = import_venues()
    print(f"-- Venues: {len(venues)} records from Venues.xlsx")
    print()
    for v in venues:
        name = escape_sql(v["name"])
        city = escape_sql(v.get("city"))
        country = escape_sql(v.get("country"))
        capacity = parse_capacity(v.get("capacity"))
        email = escape_sql(v.get("booking_email"))
        website = escape_sql(v.get("website_url"))
        source = escape_sql(v["source"])
        source_id = escape_sql(v.get("source_id"))
        raw = escape_sql(json.dumps({k: str(v2) if v2 is not None else None for k, v2 in v.items()}, ensure_ascii=False))

        print(f"INSERT INTO discovered_venues (source, source_id, name, city, country, capacity, booking_email, website_url, venue_type, raw_data)")
        print(f"VALUES ({source}, {source_id}, {name}, {city}, {country}, {capacity}, {email}, {website}, 'club', {raw})")
        print(f"ON CONFLICT DO NOTHING;")
        print()

    # Import festivals as discovered_events
    events = import_festivals()
    print(f"-- Festivals: {len(events)} records from Festivals.xlsx")
    print()
    for e in events:
        name = escape_sql(e["name"])
        city = escape_sql(e.get("city"))
        country = escape_sql(e.get("country"))
        website = escape_sql(e.get("website_url"))
        email = escape_sql(e.get("booking_email"))
        source = escape_sql(e["source"])
        source_id = escape_sql(e.get("source_id"))
        event_type = escape_sql(e.get("event_type", "festival"))
        fest_status = e.get("status", "")
        db_status = "active"
        if fest_status and "cancel" in str(fest_status).lower():
            db_status = "cancelled"
        raw = escape_sql(json.dumps({k: str(v2) if v2 is not None else None for k, v2 in e.items()}, ensure_ascii=False))

        print(f"INSERT INTO discovered_events (source, source_id, name, city, country, website_url, booking_email, event_type, status, raw_data)")
        print(f"VALUES ({source}, {source_id}, {name}, {city}, {country}, {website}, {email}, {event_type}, '{db_status}', {raw})")
        print(f"ON CONFLICT DO NOTHING;")
        print()

    # Import booking agencies/promoters as discovered_venues with venue_type set
    bookings = import_bookings()
    print(f"-- Booking agencies/promoters: {len(bookings)} records from Bookings & Promoters.xlsx")
    print()
    for b in bookings:
        name = escape_sql(b["name"])
        city = escape_sql(b.get("city"))
        country = escape_sql(b.get("country"))
        email = escape_sql(b.get("booking_email"))
        phone = escape_sql(b.get("phone"))
        website = escape_sql(b.get("website_url"))
        contact_type = b.get("contact_type", "booking_agency")
        source_id = escape_sql(b.get("source_id"))
        raw = escape_sql(json.dumps({k: str(v2) if v2 is not None else None for k, v2 in b.items()}, ensure_ascii=False))

        print(f"INSERT INTO discovered_venues (source, source_id, name, city, country, booking_email, phone, website_url, venue_type, raw_data)")
        print(f"VALUES ('excel_import', {source_id}, {name}, {city}, {country}, {email}, {phone}, {website}, '{contact_type}', {raw})")
        print(f"ON CONFLICT DO NOTHING;")
        print()

    print("COMMIT;")
    print()
    print(f"-- Total: {len(venues)} venues, {len(events)} festivals, {len(bookings)} booking contacts")

    # Summary to stderr
    # Collect country stats
    all_countries = set()
    for v in venues:
        if v.get("country"):
            all_countries.add(v["country"])
    for e in events:
        if e.get("country"):
            all_countries.add(e["country"])
    for b in bookings:
        if b.get("country"):
            all_countries.add(b["country"])

    print(f"\nImport summary:", file=sys.stderr)
    print(f"  Venues:           {len(venues)}", file=sys.stderr)
    print(f"  Festivals:        {len(events)}", file=sys.stderr)
    print(f"  Booking contacts: {len(bookings)}", file=sys.stderr)
    print(f"  Total records:    {len(venues) + len(events) + len(bookings)}", file=sys.stderr)
    print(f"  Countries:        {len(all_countries)} ({', '.join(sorted(all_countries))})", file=sys.stderr)


if __name__ == "__main__":
    main()
